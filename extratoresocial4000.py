import os
import re
import logging
import pdfplumber
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color

SETORES = {
    'operador de loja': 'Loja',
    'operador de caixa': 'Loja',
    'auxiliar de comércio': 'Loja',
    'auxiliar de loja': 'Loja',
    'encarregado': 'Estoque',
    'conferente': 'Estoque',
    'auxiliar de estoque': 'Estoque',
    'escritório': 'ADM',
    'auxiliar de Departamento Pessoal': 'ADM',
    'auxiliar de limpeza': 'Loja'
}

def extrair_dados_pdf(caminho_pdf):
    dados = {
        'cnpj': None, 'nome': None, 'rg': None, 'cpf': None,
        'nascimento': None, 'matricula': None, 'admissao': None,
        'funcao': None, 'setor': None, 'exame_medico': None
    }
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            texto = pdf.pages[0].extract_text()
            padroes = {
                'cnpj': r'C\.N\.P\.J/C\.E\.I\s*:\s*([\d./-]+)',
                'nome': r'Nome Funcionário\s*:\s*([^\n]+)',
                'rg': r'RG\s*:\s*(\d[\d.X-]*)',
                'cpf': r'CPF\s*:\s*([\d.-]+)',
                'nascimento': r'Data de Nascimento\s*:\s*(\d{2}/\d{2}/\d{4})',
                'matricula': r'Nº Registro\s*:\s*(\d+)',
                'admissao': r'Data Admissão\s*:\s*(\d{2}/\d{2}/\d{4})',
                'funcao': r'Cargo Admissão\s*:\s*([^\n]+?)(?=\s*Data exame médico|$)',
                'exame_medico': r'Data exame médico\s*:\s*(\d{2}/\d{2}/\d{4})'
            }
            for campo, padrao in padroes.items():
                match = re.search(padrao, texto)
                if match:
                    dados[campo] = match.group(1).strip()
            if dados['funcao']:
                funcao_limpa = re.sub(r'\s+', ' ', dados['funcao']).strip().lower()
                dados['setor'] = 'Outros'
                for palavra_chave, setor in SETORES.items():
                    if palavra_chave in funcao_limpa:
                        dados['setor'] = setor
                        break
            campos_faltantes = [campo for campo, valor in dados.items() if not valor]
            if campos_faltantes:
                logging.error(f"Campos faltantes em {os.path.basename(caminho_pdf)}: {', '.join(campos_faltantes)}")
            return dados
    except Exception as e:
        logging.error(f"Erro ao processar {os.path.basename(caminho_pdf)}: {str(e)}")
        return None

def criar_planilha(dados, caminho_saida):
    colunas = [
        'CNPJ DA FRENTE DE TRABALHO/ONDE ATUA',
        'NOME DO FUNCIONÁRIO',
        'RG',
        'CPF',
        'DATA DE NASCIMENTO',
        'NÚMERO DE MATRÍCULA DO ESOCIAL',
        'DATA DE ADMISSÃO',
        'DATA DE AFASTAMENTO (SE HOUVER)',
        'FUNÇÃO DO COLABORADOR',
        'SETOR QUE ATUA',
        'DATA DO ÚLTIMO EXAME REALIZADO'
    ]
    df = pd.DataFrame(dados)
    df = df.reindex(columns=colunas)
    temp_path = 'temp_planilha.xlsx'
    df.to_excel(temp_path, index=False, sheet_name='CADASTRO COLABORADORES')
    wb = Workbook()
    ws = wb.active
    ws.title = "CADASTRO COLABORADORES"
    ws.freeze_panes = "A2"
    from openpyxl.styles.colors import Color
    cor_texto = Color(rgb='0044546A')
    cor_fundo = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    estilo_cabecalho = Font(name='Calibri', size=12, bold=True, color=cor_texto)
    estilo_conteudo = Font(name='Calibri', size=11)
    estilo_texto_centralizado = Alignment(horizontal='center', vertical='center', wrap_text=True)
    estilo_texto_normal = Alignment(vertical='center')
    borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.row_dimensions[1].height = 57
    for col_idx, cabecalho in enumerate(colunas, 1):
        celula = ws.cell(row=1, column=col_idx, value=cabecalho)
        celula.font = estilo_cabecalho
        celula.fill = cor_fundo
        celula.alignment = estilo_texto_centralizado
        celula.border = borda_fina
    for linha_idx, linha in enumerate(df.values, 2):
        for col_idx, valor in enumerate(linha, 1):
            celula = ws.cell(row=linha_idx, column=col_idx, value=valor)
            celula.font = estilo_conteudo
            celula.border = borda_fina
            if col_idx != 2:
                celula.alignment = estilo_texto_centralizado
            else:
                celula.alignment = estilo_texto_normal
    for col_idx, coluna in enumerate(ws.columns, 1):
        max_length = 0
        coluna_letra = get_column_letter(col_idx)
        for celula in coluna:
            try:
                if len(str(celula.value)) > max_length:
                    max_length = len(str(celula.value))
            except:
                pass
        if col_idx == 1:
            ws.column_dimensions[coluna_letra].width = 25
        elif col_idx == 2:
            ws.column_dimensions[coluna_letra].width = max_length + 5
        else:
            ws.column_dimensions[coluna_letra].width = max(max_length + 2, 15)
    wb.save(caminho_saida)
    os.remove(temp_path)

def selecionar_arquivos():
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    arquivos = filedialog.askopenfilenames(title="Selecione as fichas de registro (PDF)", filetypes=[("Arquivos PDF", "*.pdf")])
    if not arquivos:
        return None
    caminho_saida = filedialog.asksaveasfilename(title="Salvar planilha como", defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")])
    if not caminho_saida:
        return None
    return arquivos, caminho_saida

def processar_arquivos():
    try:
        resultado = selecionar_arquivos()
        if not resultado:
            return
        arquivos_pdf, caminho_saida = resultado
        dados = []
        for pdf in arquivos_pdf:
            dados_funcionario = extrair_dados_pdf(pdf)
            if dados_funcionario:
                dados.append({
                    'CNPJ DA FRENTE DE TRABALHO/ONDE ATUA': dados_funcionario['cnpj'],
                    'NOME DO FUNCIONÁRIO': dados_funcionario['nome'],
                    'RG': dados_funcionario['rg'],
                    'CPF': dados_funcionario['cpf'],
                    'DATA DE NASCIMENTO': dados_funcionario['nascimento'],
                    'NÚMERO DE MATRÍCULA DO ESOCIAL': dados_funcionario['matricula'],
                    'DATA DE ADMISSÃO': dados_funcionario['admissao'],
                    'DATA DE AFASTAMENTO (SE HOUVER)': '',
                    'FUNÇÃO DO COLABORADOR': dados_funcionario['funcao'],
                    'SETOR QUE ATUA': dados_funcionario['setor'],
                    'DATA DO ÚLTIMO EXAME REALIZADO': dados_funcionario['exame_medico']
                })
        if not dados:
            messagebox.showwarning("Nenhum dado", "Nenhum dado válido foi extraído dos arquivos selecionados")
            return
        criar_planilha(dados, caminho_saida)
        messagebox.showinfo("Sucesso", f"Planilha gerada com sucesso:\n{caminho_saida}")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro inesperado:\n{str(e)}")
        logging.exception("Erro durante o processamento")

if __name__ == "__main__":
    processar_arquivos()
